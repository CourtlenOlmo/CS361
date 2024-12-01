from openpyxl import Workbook
import openpyxl
import os
from prettytable import PrettyTable
import zmq

if os.path.isfile("Bookkeeper.xlsx"):
      """Check if the file exists, open the file if it does"""
      workbook = openpyxl.load_workbook("Bookkeeper.xlsx")
      sheet = workbook.active
      print("test")
else:
      """If the file doesn't exist, create a new one and initialize the top row of data"""
      workbook = Workbook()
      workbook.save(filename="Bookkeeper.xlsx")
      workbook = openpyxl.load_workbook("Bookkeeper.xlsx")
      sheet = workbook.active
      a1 = sheet.cell(row=1, column=1)
      a1.value = "Author"
      a2 = sheet.cell(row=1, column=2)
      a2.value = "Title"
      a3 = sheet.cell(row=1, column=3)
      a3.value = "Rating"
      a4 = sheet.cell(row=1, column=4)
      a4.value = "Quote"
      workbook.save("Bookkeeper.xlsx")

def add_Book():
      print("\n")
      print("Please ensure that you enter these correctly, as editing is not currently implemented\n")
      title = input("Enter the name of the book: ")
      author = input("What is the name of the author: ")
      rating = input("Enter a rating for this book between 1-10: ")
      quote = input(
            "If you'd like, enter a favorite quote here, or leave it blank if you plan to input it later: ")
      if quote == 0 or None:
            quote = ""
      confirmation = input(f"Would you like to enter {title} by {author}? Enter Y/N: ")
      if confirmation == "n" or confirmation == "N":
            return()

      entry = (author, title, rating, quote)

      sheet.append(entry)
      workbook.save("BookKeeper.xlsx")
      print("Your book has been added!")
      print("\n")
      return()

def see_Books():
      print("/n")

      """Create the table to be used"""
      myTable = PrettyTable(["Author", "Title", "Rating", "Quotes"])
      max_row = str(sheet.max_row)
      column = sheet.max_column
      cell = sheet["A1":"D"+max_row]
      data = [list(elem) for elem in cell]
      cell = []
      row = 2
      column = 1

      """Loop throught the list of cells and add the rows to the table"""
      for i in data:
            for n in range(1, 5):
                  cell.append(sheet.cell(row, column = n))
            if column < 4:
                  column += 1
            row += 1
            if row <= int(max_row)+1:
                  myTable.add_row([cell[0].value,cell[1].value,cell[2].value,cell[3].value])
                  cell = []
      print(myTable)
      print("\n")
      return()

def lookUp_Book():
    context = zmq.Context()
    print(">>> Client attempting to connect to server...")
    socket = context.socket(zmq.REQ)
    socket.connect(f"tcp://localhost:5555")

    while True:
        input_string = input("Enter a book to lookup, or enter Q to quit: ")
        if input_string.upper() == "Q":
            print(f">>> Exiting Program...")
            context.destroy()
            break
        else:
            print(f">>> Sending request...")
            socket.send_string(input_string)
            message = socket.recv()
            print(f">>> Server sent back: {message.decode()}")
    context.destroy()

def average_Rating():
      column = 3
      maxRow = sheet.max_row
      row = 2
      ratings = []
      ratingSum = 0
      for i in range(2, maxRow+1):
            ratings.append(int(sheet.cell(i, column).value))
      ratingSum = sum(ratings)
      ratingAverage = round(ratingSum / (maxRow - 1), 2)
      print(f"The average rating across your books is {ratingAverage}")
      print("\n")
      return()

def add_goal():
      context = zmq.Context()
      print(">>> Client attempting to connect to server...")
      socket = context.socket(zmq.REQ)
      socket.connect(f"tcp://localhost:5556")

      while True:
            input_string = input("enter the amount of hours you would like to read this week: ")
            print(f">>> Sending request...")
            socket.send_string(input_string)
            break
      context.destroy()

def main_Menu():
      selection = ""
      while selection != 0:
            print("Welcome to Bookkeeper, this program will help you track the books you've read by storing them in an excel document\n\n"
                  "Enter 1 to add a new book\n"
                  "Enter 2 to see the books in your collection\n"
                  "Enter 3 see the average rating of your books\n"
                  "Enter 4 to find an amazon link for a book\n"
                  "Enter 5 to add a weekly goal to your text file\n"
                  "Enter 0 to exit the program\n"
                  "Please only enter numbers between 0-3\n")
            selection = input("Enter your selection here: ")

            if selection == "0":
                  exit()
            if selection == "1":
                  add_Book()
            elif selection == "2":
                  see_Books()
            elif selection == "3":
                  average_Rating()
            elif selection == "4":
                  lookUp_Book()
            elif selection == "5":
                  add_goal()
      exit()

if __name__ == "__main__":
      main_Menu()


